# -*- coding: utf-8 -*-
import serial
import serial.tools.list_ports
import sys
import math
from datetime import datetime
from openpyxl import Workbook
import re
import socket
import base64
from PyQt5.QtCore import *
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
import pynmea2
def extract_lat_lon_alt(nmea_sentence):
    msg = pynmea2.parse(nmea_sentence)
    if isinstance(msg, pynmea2.GGA):
        return msg.latitude, msg.longitude, msg.altitude
    return None, None
def extract_speed_kmh(nmea_sentence):
    msg = pynmea2.parse(nmea_sentence)
    if isinstance(msg, pynmea2.RMC):
        speed_kmh = msg.spd_over_grnd * 1.852
        return speed_kmh
    return None
class information_show_screen(object):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(1022, 595)
        self.columnView = QtWidgets.QColumnView(Form)
        self.columnView.setGeometry(QtCore.QRect(10, 100, 400, 120))
        self.columnView.setStyleSheet("background: qlineargradient(x1:0, y1:0, x2:1, y2:0,\n"
"                                stop:0 rgba(89, 217, 212, 255),\n"
"                                stop:1 rgba(69, 130, 191, 255));\n"
"border-radius:30px;\n"
"\n"
"")
        self.columnView.setObjectName("columnView")
        self.label_2 = QtWidgets.QLabel(Form)
        self.label_2.setGeometry(QtCore.QRect(20, 110, 170, 40))
        font = QtGui.QFont()
        font.setFamily("Samyak Devanagari")
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.lcdNumber = QtWidgets.QLCDNumber(Form)
        self.lcdNumber.setGeometry(QtCore.QRect(170, 170, 200, 40))
        self.lcdNumber.setObjectName("lcdNumber")
        self.label = QtWidgets.QLabel(Form)
        self.label.setGeometry(QtCore.QRect(0, 0, 1024, 600))
        font = QtGui.QFont()
        font.setPointSize(8)
        self.label.setFont(font)
        self.label.setStyleSheet("background-color: rgba(89, 217, 212,50);  ")
        self.label.setText("")
        self.label.setObjectName("label")
        self.columnView_2 = QtWidgets.QColumnView(Form)
        self.columnView_2.setGeometry(QtCore.QRect(10, 280, 400, 120))
        self.columnView_2.setStyleSheet("background: qlineargradient(x1:0, y1:0, x2:1, y2:0,\n"
"                                stop:0 rgba(89, 217, 212, 255),\n"
"                                stop:1 rgba(69, 130, 191, 255));\n"
"border-radius:30px;\n"
"\n"
"")
        self.columnView_2.setObjectName("columnView_2")
        self.lcdNumber_2 = QtWidgets.QLCDNumber(Form)
        self.lcdNumber_2.setGeometry(QtCore.QRect(170, 350, 200, 40))
        self.lcdNumber_2.setObjectName("lcdNumber_2")
        self.label_3 = QtWidgets.QLabel(Form)
        self.label_3.setGeometry(QtCore.QRect(20, 290, 170, 40))
        font = QtGui.QFont()
        font.setFamily("Samyak Devanagari")
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.lcdNumber_3 = QtWidgets.QLCDNumber(Form)
        self.lcdNumber_3.setGeometry(QtCore.QRect(170, 530, 200, 40))
        self.lcdNumber_3.setObjectName("lcdNumber_3")
        self.columnView_3 = QtWidgets.QColumnView(Form)
        self.columnView_3.setGeometry(QtCore.QRect(10, 460, 400, 120))
        self.columnView_3.setStyleSheet("background: qlineargradient(x1:0, y1:0, x2:1, y2:0,\n"
"                                stop:0 rgba(89, 217, 212, 255),\n"
"                                stop:1 rgba(69, 130, 191, 255));\n"
"border-radius:30px;\n"
"\n"
"")
        self.columnView_3.setObjectName("columnView_3")
        self.label_4 = QtWidgets.QLabel(Form)
        self.label_4.setGeometry(QtCore.QRect(20, 470, 170, 40))
        font = QtGui.QFont()
        font.setFamily("Samyak Devanagari")
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.pushButton = QtWidgets.QPushButton(Form)
        self.pushButton.setGeometry(QtCore.QRect(450, 100, 120, 50))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.pushButton.setFont(font)
        self.pushButton.setIconSize(QtCore.QSize(8, 8))
        self.pushButton.setObjectName("pushButton")
        self.label_5 = QtWidgets.QLabel(Form)
        self.label_5.setGeometry(QtCore.QRect(330, 10, 350, 50))
        font = QtGui.QFont()
        font.setFamily("Ubuntu Condensed")
        font.setPointSize(28)
        font.setBold(True)
        font.setWeight(75)
        self.label_5.setFont(font)
        self.label_5.setAlignment(QtCore.Qt.AlignCenter)
        self.label_5.setObjectName("label_5")
        self.pushButton_2 = QtWidgets.QPushButton(Form)
        self.pushButton_2.setGeometry(QtCore.QRect(450, 170, 120, 50))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setIconSize(QtCore.QSize(8, 8))
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_3 = QtWidgets.QPushButton(Form)
        self.pushButton_3.setGeometry(QtCore.QRect(450, 350, 120, 50))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.pushButton_3.setFont(font)
        self.pushButton_3.setIconSize(QtCore.QSize(8, 8))
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_4 = QtWidgets.QPushButton(Form)
        self.pushButton_4.setGeometry(QtCore.QRect(450, 280, 120, 50))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.pushButton_4.setFont(font)
        self.pushButton_4.setIconSize(QtCore.QSize(8, 8))
        self.pushButton_4.setObjectName("pushButton_4")
        self.pushButton_5 = QtWidgets.QPushButton(Form)
        self.pushButton_5.setGeometry(QtCore.QRect(450, 530, 120, 50))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.pushButton_5.setFont(font)
        self.pushButton_5.setIconSize(QtCore.QSize(8, 8))
        self.pushButton_5.setObjectName("pushButton_5")
        self.pushButton_6 = QtWidgets.QPushButton(Form)
        self.pushButton_6.setGeometry(QtCore.QRect(450, 460, 120, 50))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.pushButton_6.setFont(font)
        self.pushButton_6.setIconSize(QtCore.QSize(8, 8))
        self.pushButton_6.setObjectName("pushButton_6")
        self.pushButton_7 = QtWidgets.QPushButton(Form)
        self.pushButton_7.setGeometry(QtCore.QRect(900, 20, 100, 40))
        self.pushButton_7.setObjectName("pushButton_7")
        self.columnView_4 = QtWidgets.QColumnView(Form)
        self.columnView_4.setGeometry(QtCore.QRect(610, 100, 400, 120))
        self.columnView_4.setStyleSheet("background: qlineargradient(x1:0, y1:0, x2:1, y2:0,\n"
"                                stop:0 rgba(89, 217, 212, 255),\n"
"                                stop:1 rgba(69, 130, 191, 255));\n"
"border-radius:30px;\n"
"\n"
"")
        self.columnView_4.setObjectName("columnView_4")
        self.lcdNumber_4 = QtWidgets.QLCDNumber(Form)
        self.lcdNumber_4.setGeometry(QtCore.QRect(770, 170, 200, 40))
        self.lcdNumber_4.setObjectName("lcdNumber_4")
        self.label_6 = QtWidgets.QLabel(Form)
        self.label_6.setGeometry(QtCore.QRect(620, 110, 170, 40))
        font = QtGui.QFont()
        font.setFamily("Samyak Devanagari")
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.columnView_5 = QtWidgets.QColumnView(Form)
        self.columnView_5.setGeometry(QtCore.QRect(610, 280, 400, 120))
        self.columnView_5.setStyleSheet("background: qlineargradient(x1:0, y1:0, x2:1, y2:0,\n"
"                                stop:0 rgba(89, 217, 212, 255),\n"
"                                stop:1 rgba(69, 130, 191, 255));\n"
"border-radius:30px;\n"
"\n"
"")
        self.columnView_5.setObjectName("columnView_5")
        self.lcdNumber_5 = QtWidgets.QLCDNumber(Form)
        self.lcdNumber_5.setGeometry(QtCore.QRect(770, 350, 200, 40))
        self.lcdNumber_5.setObjectName("lcdNumber_5")
        self.label_7 = QtWidgets.QLabel(Form)
        self.label_7.setGeometry(QtCore.QRect(620, 290, 170, 40))
        font = QtGui.QFont()
        font.setFamily("Samyak Devanagari")
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.label_8 = QtWidgets.QLabel(Form)
        self.label_8.setGeometry(QtCore.QRect(620, 470, 170, 40))
        font = QtGui.QFont()
        font.setFamily("Samyak Devanagari")
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.label_8.setFont(font)
        self.label_8.setObjectName("label_8")
        self.lcdNumber_6 = QtWidgets.QLCDNumber(Form)
        self.lcdNumber_6.setGeometry(QtCore.QRect(770, 530, 200, 40))
        self.lcdNumber_6.setObjectName("lcdNumber_6")
        self.columnView_6 = QtWidgets.QColumnView(Form)
        self.columnView_6.setGeometry(QtCore.QRect(610, 460, 400, 120))
        self.columnView_6.setStyleSheet("background: qlineargradient(x1:0, y1:0, x2:1, y2:0,\n"
"                                stop:0 rgba(89, 217, 212, 255),\n"
"                                stop:1 rgba(69, 130, 191, 255));\n"
"border-radius:30px;\n"
"\n"
"")
        self.columnView_6.setObjectName("columnView_6")
        self.label.raise_()
        self.columnView_6.raise_()
        self.columnView_3.raise_()
        self.columnView.raise_()
        self.columnView_2.raise_()
        self.label_2.raise_()
        self.lcdNumber.raise_()
        self.lcdNumber_2.raise_()
        self.label_3.raise_()
        self.lcdNumber_3.raise_()
        self.label_4.raise_()
        self.pushButton.raise_()
        self.label_5.raise_()
        self.pushButton_2.raise_()
        self.pushButton_3.raise_()
        self.pushButton_4.raise_()
        self.pushButton_5.raise_()
        self.pushButton_6.raise_()
        self.pushButton_7.raise_()
        self.columnView_4.raise_()
        self.lcdNumber_4.raise_()
        self.label_6.raise_()
        self.columnView_5.raise_()
        self.lcdNumber_5.raise_()
        self.label_7.raise_()
        self.label_8.raise_()
        self.lcdNumber_6.raise_()
        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)
    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Form"))
        self.label_2.setText(_translate("Form", "码盘速度显示"))
        self.label_3.setText(_translate("Form", "雷达速度显示"))
        self.label_4.setText(_translate("Form", "GNSS速度显示"))
        self.pushButton.setText(_translate("Form", "打开端口"))
        self.label_5.setText(_translate("Form", "作业数据采集系统"))
        self.pushButton_2.setText(_translate("Form", "关闭端口"))
        self.pushButton_3.setText(_translate("Form", "关闭端口"))
        self.pushButton_4.setText(_translate("Form", "打开端口"))
        self.pushButton_5.setText(_translate("Form", "关闭端口"))
        self.pushButton_6.setText(_translate("Form", "打开端口"))
        self.pushButton_7.setText(_translate("Form", "返回"))
        self.label_6.setText(_translate("Form", "角度显示"))
        self.label_7.setText(_translate("Form", "力F1显示"))
        self.label_8.setText(_translate("Form", "力F2显示"))
class WelcomeScreen(object):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(1024, 600)
        self.pushButton = QtWidgets.QPushButton(Form)
        self.pushButton.setGeometry(QtCore.QRect(590, 420, 300, 60))
        font = QtGui.QFont()
        font.setFamily("楷体")
        font.setPointSize(30)
        font.setBold(True)
        font.setUnderline(False)
        font.setWeight(75)
        self.pushButton.setFont(font)
        self.pushButton.setStyleSheet("QPushButton:hover {\n"
"    border: 2px solid green;\n"
"}\n"
"QPushButton {\n"
"    border: none;\n"
"}\n"
"\n"
"\n"
"")
        self.pushButton.setObjectName("pushButton")
        self.label = QtWidgets.QLabel(Form)
        self.label.setGeometry(QtCore.QRect(0, 0, 1024, 600))
        self.label.setText("")
        self.label.setPixmap(QtGui.QPixmap("icon/ResizedImage.png"))
        self.label.setObjectName("label")
        self.textEdit = QtWidgets.QTextEdit(Form)
        self.textEdit.setGeometry(QtCore.QRect(520, 190, 421, 191))
        font = QtGui.QFont()
        font.setFamily("楷体")
        self.textEdit.setFont(font)
        self.textEdit.setStyleSheet("background: transparent;\n"
"border: none;\n"
"")
        self.textEdit.setObjectName("textEdit")
        self.label.raise_()
        self.textEdit.raise_()
        self.pushButton.raise_()

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Form"))
        self.pushButton.setText(_translate("Form", "点击继续"))
        self.textEdit.setMarkdown(_translate("Form", "**果园作业机械研究团队**\n"
"\n"
""))
        self.textEdit.setHtml(_translate("Form", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'楷体\'; font-size:11pt; font-weight:400; font-style:normal;\">\n"
"<p align=\"center\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:45pt; font-weight:600;\">果园作业机械研究团队</span></p></body></html>"))



class WelcomeWindow(QWidget):
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.ui = WelcomeScreen()
        self.ui.setupUi(self)
        self.setFixedSize(self.size())
        self.ui.pushButton.clicked.connect(self.switch_window)
    def switch_window(self):
        self.stacked_widget.setCurrentIndex(1)
def handle_collected_gnss_data(speed_data, time_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Speed Data"
    ws.append(["Time Stamp", "Speed"])
    for time_stamp, speed in zip(time_data, speed_data):
        ws.append([time_stamp.strftime("%Y-%m-%d %H:%M:%S"), speed])
    file_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + "_gnss_speed.xlsx"
    wb.save(file_name)
    print(f"Data saved to {file_name}")
def handle_collected_wheel_data(speed_data, time_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Speed Data"
    ws.append(["Time Stamp", "Speed"])
    for time_stamp, speed in zip(time_data, speed_data):
        ws.append([time_stamp.strftime("%Y-%m-%d %H:%M:%S"), speed])
    file_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + "_wheel_speed.xlsx"
    wb.save(file_name)
    print(f"Data saved to {file_name}")
def handle_collected_radar_data(speed_data, angle_data, force1_data, force2_data, time_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Radar Data"
    ws.append(["Time Stamp", "Speed", "Angle", "Force1", "Force2"])
    for time_stamp, speed, angle, force1, force2 in zip(time_data, speed_data, angle_data,
                                                        force1_data, force2_data):
        ws.append([time_stamp.strftime("%Y-%m-%d %H:%M:%S"), speed, angle, force1, force2])
    file_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + "_radar_data.xlsx"
    wb.save(file_name)
    print(f"Data saved to {file_name}")
def handle_collected_acc_data(accX_list, accY_list, accZ_list, time_stamps):
    print("保存IMU数据")
    wb = Workbook()
    ws = wb.active
    ws.title = "IMU Data"
    columns = ["Time Stamp", "AccX", "AccY", "AccZ"]
    ws.append(columns)
    for i in range(len(time_stamps)):
        time_stamp = time_stamps[i].strftime("%Y-%m-%d %H:%M:%S")
        accX = accX_list[i]
        accY = accY_list[i]
        accZ = accZ_list[i]
        ws.append([time_stamp, accX, accY, accZ])
    file_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + "_acc_data.xlsx"
    wb.save(file_name)
    print(f"Data saved to {file_name}")
def handle_collected_gyr_data(gyroX_list, gyroY_list, gyroZ_list, time_stamps):
    wb = Workbook()
    ws = wb.active
    ws.title = "IMU Data"
    columns = ["Time Stamp", "AccX", "AccY", "AccZ"]
    ws.append(columns)
    for i in range(len(time_stamps)):
        time_stamp = time_stamps[i].strftime("%Y-%m-%d %H:%M:%S")
        gyroX = gyroX_list[i]
        gyroY = gyroY_list[i]
        gyroZ = gyroZ_list[i]
        ws.append([time_stamp, gyroX, gyroY, gyroZ])
    file_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + "_gyro_data.xlsx"
    wb.save(file_name)
    print(f"Data saved to {file_name}")
def handle_collected_ord_data(roll_list, pitch_list, yaw_list, time_stamps):
    print("保存IMU数据")
    wb = Workbook()
    ws = wb.active
    ws.title = "IMU Data"
    columns = ["Time Stamp", "roll", "pitch", "yaw"]
    ws.append(columns)
    for i in range(len(time_stamps)):
        time_stamp = time_stamps[i].strftime("%Y-%m-%d %H:%M:%S")
        roll = roll_list[i]
        pitch = pitch_list[i]
        yaw = yaw_list[i]
        ws.append([time_stamp, roll, pitch, yaw])
    file_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + "_ord_data.xlsx"
    wb.save(file_name)
    print(f"Data saved to {file_name}")
class information_show_Window(QWidget):
    def __init__(self, stacked_widget):
        super().__init__()
        self.gnss_work = None
        self.radar_work = None
        self.wheel_work = None
        self.ui = information_show_screen()
        self.ui.setupUi(self)
        self.ui_process()
        self.port = None
        self.baudrate = None
        self.stacked_widget = stacked_widget
        self.speed_data = []
        self.speed_data_wheel = []
        self.speed_data_radar = []
    def ui_process(self):
        self.speed_show_lcd_for_wheel = self.ui.lcdNumber
        self.speed_show_lcd_for_radar = self.ui.lcdNumber_2
        self.speed_show_lcd_for_GNSS = self.ui.lcdNumber_3
        self.angle_show_lcd = self.ui.lcdNumber_4
        self.force_show_lcd_for_F1 = self.ui.lcdNumber_5
        self.force_show_lcd_for_F2 = self.ui.lcdNumber_6
        self.open_speed_show_for_wheel = self.ui.pushButton
        self.close_speed_show_for_wheel = self.ui.pushButton_2
        self.open_speed_show_for_radar = self.ui.pushButton_4
        self.close_speed_show_for_radar = self.ui.pushButton_3
        self.open_speed_show_for_gnss = self.ui.pushButton_6
        self.close_speed_show_for_gnss = self.ui.pushButton_5
        self.open_speed_show_for_wheel.clicked.connect(self.wheel_choose_port)
        self.close_speed_show_for_wheel.clicked.connect(self.wheel_speed_fun_end)
        self.open_speed_show_for_radar.clicked.connect(self.radar_choose_port)
        self.close_speed_show_for_radar.clicked.connect(self.radar_speed_fun_end)
        self.open_speed_show_for_gnss.clicked.connect(self.gnss_choose_port)
        self.close_speed_show_for_gnss.clicked.connect(self.gnss_speed_fun_end)
        self.back_menu = self.ui.pushButton_7
        self.back_menu.clicked.connect(self.return_to_welcome_screen)
    def return_to_welcome_screen(self):
        self.stacked_widget.setCurrentIndex(0)
    def wheel_choose_port(self):
        self.dialog_wheel = PortSelectionDialog_wheel()
        self.dialog_wheel.port_baudrate_selected_signal.connect(self.whell_port_baudrate_selected)
        self.dialog_wheel.exec_()
    def whell_port_baudrate_selected(self, chosen_port, chosen_baudrate):
        try:
            self.now = datetime.now()
            self.wheel_work = Worker_wheel(chosen_port, chosen_baudrate)
            self.wheel_work.finished_signal.connect(self.wheel_worker_finished)
            self.wheel_work.data_collected_signal.connect(handle_collected_wheel_data)
            self.wheel_work.tran_speed_signal.connect(self.speed_wheel_show)
            self.wheel_work.mistake_message_transmit.connect(self.mistake_message_show)
            self.wheel_work.start()
        except serial.serialutil.SerialException as error:
            self.mistake_message_show(str(error))
    def wheel_speed_fun_end(self):
        if self.wheel_work is not None:
            self.wheel_work.should_run = False
            self.wheel_work.quit()
            self.wheel_work.wait()
    def wheel_worker_finished(self):
        self.wheel_work = None
    def speed_wheel_show(self, speed):
        self.speed_show_lcd_for_wheel.display(speed)
        self.speed_show_lcd_for_wheel.show()
    def radar_choose_port(self):
        self.dialog_radar = PortSelectionDialog_radar()
        self.dialog_radar.port_baudrate_selected_signal.connect(self.radar_port_baudrate_selected)
        self.dialog_radar.exec_()
    def radar_port_baudrate_selected(self, chosen_port, chosen_baudrate):
        try:
            self.radar_now = datetime.now()
            self.radar_work = Worker_radar(chosen_port, chosen_baudrate)
            self.radar_work.finished_signal.connect(self.radar_worker_finished)
            self.radar_work.data_collected_signal.connect(handle_collected_radar_data)
            self.radar_work.acc_data_collected_signal.connect(handle_collected_acc_data)
            self.radar_work.gyr_data_collected_signal.connect(handle_collected_gyr_data)
            self.radar_work.ord_data_collected_signal.connect(handle_collected_ord_data)
            self.radar_work.avg_data_parsed.connect(self.radar_data_show)
            self.radar_work.mistake_message_transmit.connect(self.mistake_message_show)
            self.radar_work.start()
        except serial.serialutil.SerialException as error:
            self.mistake_message_show(error)
    def radar_speed_fun_end(self):
        if self.radar_work is not None:
            self.radar_work.should_run = False
            self.radar_work.quit()
            self.radar_work.wait()
    def radar_worker_finished(self):
        self.radar_work = None
    def radar_data_show(self, speed, angle, force1, force2):
        self.speed_show_lcd_for_radar.display(speed)
        self.speed_show_lcd_for_radar.show()
        self.angle_show_lcd.display(angle)
        self.angle_show_lcd.show()
        self.force_show_lcd_for_F1.display(force1)
        self.force_show_lcd_for_F1.show()
        self.force_show_lcd_for_F2.display(force2)
        self.force_show_lcd_for_F2.show()
    def gnss_choose_port(self):
        self.dialog_gnss = PortSelectionDialog_gnss()
        self.dialog_gnss.port_baudrate_ntrip_details_signal.connect(self.gnss_port_baudrate_ntrip_details_selected)
        self.dialog_gnss.exec_()
    def gnss_port_baudrate_ntrip_details_selected(self, chosen_port, chosen_baudrate, ntrip_host, ntrip_port,
                                                  ntrip_user, ntrip_password, ntrip_mountpoint):
        try:
            self.gnss_work = Worker_gnss(chosen_port, chosen_baudrate, ntrip_host, ntrip_port, ntrip_user,
                                         ntrip_password, ntrip_mountpoint)
            self.gnss_work.finished_signal.connect(self.gnss_worker_finished)
            self.gnss_work.average_speed_signal.connect(self.speed_gnss_show)
            self.gnss_work.data_collected_signal.connect(handle_collected_gnss_data)
            self.gnss_work.mistake_message_transmit.connect(self.mistake_message_show)
            self.gnss_work.start()
        except serial.serialutil.SerialException as error:
            self.mistake_message_show(error)
    def gnss_speed_fun_end(self):
        if self.gnss_work is not None:
            self.gnss_work.should_run = False
            self.gnss_work.quit()
            self.gnss_work.wait()
    def gnss_worker_finished(self):
        self.gnss_work = None
    def speed_gnss_show(self, speed):
        self.speed_show_lcd_for_GNSS.display(speed)
        self.speed_show_lcd_for_GNSS.show()
    def mistake_message_show(self, message):
        QMessageBox.information(self, "错误", "错误原因{}".format(message))
class PortSelectionDialog_wheel(QDialog):
    port_baudrate_selected_signal = pyqtSignal(str, int)
    def __init__(self):
        super().__init__()
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        layout = QVBoxLayout(self)
        self.comboBoxPort = QComboBox()
        self.populate_ports()
        layout.addWidget(self.comboBoxPort)
        self.comboBoxBaudrate = QComboBox()
        self.comboBoxBaudrate.addItems(["9600", "19200", "38400", "57600", "115200"])  # 常用波特率
        layout.addWidget(self.comboBoxBaudrate)
        self.confirmButton = QPushButton("确认")
        self.confirmButton.clicked.connect(self.confirm)
        layout.addWidget(self.confirmButton)
        self.closeButton = QPushButton("关闭")
        self.closeButton.clicked.connect(self.close)
        layout.addWidget(self.closeButton)
    def confirm(self):
        chosen_port = self.comboBoxPort.currentText()
        chosen_baudrate = int(self.comboBoxBaudrate.currentText())
        self.port_baudrate_selected_signal.emit(chosen_port, chosen_baudrate)
        self.accept()
    def populate_ports(self):
        self.comboBoxPort.clear()
        ports = serial.tools.list_ports.comports()
        available_ports = [port.device for port in ports]
        self.comboBoxPort.addItems(available_ports)
class PortSelectionDialog_radar(QDialog):
    port_baudrate_selected_signal = pyqtSignal(str, int)
    def __init__(self):
        super().__init__()
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        layout = QVBoxLayout(self)
        self.comboBoxPort = QComboBox()
        self.populate_ports()
        layout.addWidget(self.comboBoxPort)
        self.comboBoxBaudrate = QComboBox()
        self.comboBoxBaudrate.addItems(["9600", "19200", "38400", "57600", "115200"])  # 常用波特率
        layout.addWidget(self.comboBoxBaudrate)
        self.confirmButton = QPushButton("确认")
        self.confirmButton.clicked.connect(self.confirm)
        layout.addWidget(self.confirmButton)
        self.closeButton = QPushButton("关闭")
        self.closeButton.clicked.connect(self.close)
        layout.addWidget(self.closeButton)
    def confirm(self):
        chosen_port = self.comboBoxPort.currentText()
        chosen_baudrate = int(self.comboBoxBaudrate.currentText())
        self.port_baudrate_selected_signal.emit(chosen_port, chosen_baudrate)
        self.accept()
    def populate_ports(self):
        self.comboBoxPort.clear()
        ports = serial.tools.list_ports.comports()
        available_ports = [port.device for port in ports]
        self.comboBoxPort.addItems(available_ports)
class PortSelectionDialog_gnss(QDialog):
    port_baudrate_ntrip_details_signal = pyqtSignal(str, int, str, str, str, str, str)
    def __init__(self):
        super().__init__()
        self.init_ui()
    def init_ui(self):
        self.setWindowTitle("选择端口和波特率")
        layout = QVBoxLayout(self)
        self.comboBoxPort = QComboBox()
        self.populate_ports()
        layout.addWidget(self.comboBoxPort)
        self.comboBoxBaudrate = QComboBox()
        self.comboBoxBaudrate.addItems(["115200"])
        layout.addWidget(self.comboBoxBaudrate)
        self.comboBoxRTKUse = QComboBox()
        self.comboBoxRTKUse.addItems(["不使用RTK", "测试账户", "用户账户"])
        layout.addWidget(self.comboBoxRTKUse)
        self.comboBoxRTKUse.currentIndexChanged.connect(self.onRTKUseChanged)  # 连接变更事件
        self.ntripHostInput = QLineEdit()
        layout.addWidget(self.ntripHostInput)
        self.ntripPortInput = QLineEdit()
        layout.addWidget(self.ntripPortInput)
        self.ntripUserInput = QLineEdit()
        layout.addWidget(self.ntripUserInput)
        self.ntripPasswordInput = QLineEdit()
        layout.addWidget(self.ntripPasswordInput)
        self.ntripMountpointInput = QLineEdit()
        layout.addWidget(self.ntripMountpointInput)
        self.confirmButton = QPushButton("确认")
        self.confirmButton.clicked.connect(self.confirm)
        layout.addWidget(self.confirmButton)
        self.closeButton = QPushButton("关闭")
        self.closeButton.clicked.connect(self.close)
        layout.addWidget(self.closeButton)
        self.onRTKUseChanged()
    def onRTKUseChanged(self):
        index = self.comboBoxRTKUse.currentIndex()
        if index == 0:
            self.setAutoFillNtripSettings("0")
        elif index == 1:
            self.setAutoFillNtripSettings("1")
        else:
            self.setAutoFillNtripSettings("", allowEdit=True)
    def setAutoFillNtripSettings(self, value, allowEdit=False):
        self.ntripHostInput.setText(value)
        self.ntripPortInput.setText(value)
        self.ntripUserInput.setText(value)
        self.ntripPasswordInput.setText(value)
        self.ntripMountpointInput.setText(value)
        self.ntripHostInput.setReadOnly(not allowEdit)
        self.ntripPortInput.setReadOnly(not allowEdit)
        self.ntripUserInput.setReadOnly(not allowEdit)
        self.ntripPasswordInput.setReadOnly(not allowEdit)
        self.ntripMountpointInput.setReadOnly(not allowEdit)
    def confirm(self):
        chosen_port = self.comboBoxPort.currentText()
        chosen_baudrate = int(self.comboBoxBaudrate.currentText())
        ntrip_host = self.ntripHostInput.text()
        ntrip_port = self.ntripPortInput.text()
        ntrip_user = self.ntripUserInput.text()
        ntrip_password = self.ntripPasswordInput.text()
        ntrip_mountpoint = self.ntripMountpointInput.text()
        if not (ntrip_host and ntrip_port and ntrip_user and ntrip_password and ntrip_mountpoint):
            QMessageBox.warning(self, "输入错误", "有变量输入为空，请检查。")
            return
        print(chosen_port, chosen_baudrate, ntrip_host, ntrip_port, ntrip_user,
              ntrip_password, ntrip_mountpoint)
        self.port_baudrate_ntrip_details_signal.emit(chosen_port, chosen_baudrate, ntrip_host, ntrip_port, ntrip_user,
                                                     ntrip_password, ntrip_mountpoint)
        self.accept()
    def populate_ports(self):
        ports = serial.tools.list_ports.comports()
        port_list = [port.device for port in ports]
        self.comboBoxPort.addItems(port_list)
class Worker_wheel(QThread):
    finished_signal = pyqtSignal()
    tran_speed_signal = pyqtSignal(float)
    mistake_message_transmit = pyqtSignal(str)
    data_collected_signal = pyqtSignal(list, list)
    def __init__(self, port, baudrate):
        super(Worker_wheel, self).__init__()
        self.should_run = True
        self.port = port
        self.baudrate = baudrate
        self.speed_data = []
        self.time_stamps = []
        self.speed_data_for_avg = []
    def run(self):
        try:
            ser = serial.Serial(self.port, self.baudrate, timeout=1)
            while self.should_run:
                try:
                    data_line = ser.readline().decode('utf-8').rstrip()
                    if "Velocity = " in data_line:
                        velocity_str = data_line.replace("Velocity = ", "")
                        speed = float(velocity_str)
                        self.speed_data.append(speed)
                        self.time_stamps.append(datetime.now())
                        self.tran_speed_signal.emit(speed)
                except ValueError:
                    continue
        except Exception as err_1:
            self.mistake_message_transmit.emit(str(err_1))
        finally:
            self.data_collected_signal.emit(self.speed_data, self.time_stamps)
            self.finished_signal.emit()
class Worker_gnss(QThread):
    finished_signal = pyqtSignal()
    mistake_message_transmit = pyqtSignal(str)
    average_speed_signal = pyqtSignal(float)
    data_collected_signal = pyqtSignal(list, list)
    def __init__(self, port, baudrate, ntrip_host, ntrip_port, ntrip_user, ntrip_password, ntrip_mountpoint):
        super(Worker_gnss, self).__init__()
        self.RTK_singal = 0
        self.point = 0
        self.baudrate = int(baudrate)
        self.port = port
        self.ntrip_host = ntrip_host
        self.ntrip_port = ntrip_port
        self.ntrip_user = ntrip_user
        self.ntrip_password = ntrip_password
        self.ntrip_mountpoint = ntrip_mountpoint
        self.should_run = True
        self.speed_data = []
        self.time_stamps = []
        self.speed_data_for_avg = []
        self.Ntrip_host = "203.107.45.154"
        self.Ntrip_port = 8002
        self.Ntrip_mountpoint = "AUTO"
        self.Ntrip_user = "qxvnzk001057"
        self.Ntrip_password = "f2966e2"
    def run(self):
        try:
            lat, lon, alt, speed, rtk_fix_quality = 0, 0, 0, 0.0, 0
            gngga_messages = []
            if self.baudrate is not None and self.port is not None:
                ser = serial.Serial(self.port, self.baudrate, timeout=1)
                ser.write("gnrmc 0.1\r\n".encode())
                ser.write("gngga 0.1\r\n".encode())
                self.reconnect()
                while self.should_run:
                    try:
                        if self.ntrip_host == "0":
                            raw_message = ser.readline()
                            gnss_message = raw_message.decode().strip()
                            if gnss_message.startswith("$GNGGA"):
                                lat, lon, alt = extract_lat_lon_alt(gnss_message)
                            raw_message = ser.readline()
                            gnss_message = raw_message.decode().strip()
                            if gnss_message.startswith("$GNRMC"):
                                speed = extract_speed_kmh(gnss_message)
                            self.speed_data.append(speed)
                            self.speed_data_for_avg.append(speed)
                            self.time_stamps.append(datetime.now())
                            if len(self.speed_data_for_avg) == 10:
                                avg_speed = sum(self.speed_data_for_avg) / 10
                                self.average_speed_signal.emit(avg_speed)
                                self.speed_data_for_avg.clear()
                        elif self.ntrip_host == "1" and self.RTK_singal == 1:
                            self.point = 1
                            raw_message_gngga = ser.readline()
                            gngga_message = raw_message_gngga.decode().strip()
                            if gngga_message.startswith("$GNGGA"):
                                fields = gngga_message.split(",")
                                rtk_fix_quality = fields[6]
                                gngga_messages.append(gngga_message)
                                if len(gngga_messages) % 20 == 0:
                                    gnss_messages_send = "\r\n".join(gngga_messages) + "\r\n"
                                    print(gnss_messages_send)
                                    self.s.send(gnss_messages_send.encode())
                                    gngga_messages = []
                                    rtcm_data = self.s.recv(102400)
                                    if rtcm_data is not None:
                                        ser.write(rtcm_data)
                                        ser.flush()
                            raw_message_gnrmc = ser.readline()
                            gnrmc_message = raw_message_gnrmc.decode().strip()
                            print(gnrmc_message)
                            if gnrmc_message.startswith("$GNRMC"):
                                print(rtk_fix_quality)
                                if rtk_fix_quality == '4' or rtk_fix_quality == '5':
                                    speed = extract_speed_kmh(gnrmc_message)
                                    print(speed)
                                else:
                                    speed = 0
                            self.speed_data.append(speed)
                            self.speed_data_for_avg.append(speed)
                            self.time_stamps.append(datetime.now())
                            if len(self.speed_data_for_avg) == 10:
                                avg_speed = sum(self.speed_data_for_avg) / 10
                                self.average_speed_signal.emit(avg_speed)
                                self.speed_data_for_avg.clear()
                            ser.flushInput()
                            ser.flushOutput()
                    except ValueError:
                        continue
        except Exception as err_1:
            self.mistake_message_transmit.emit(str(err_1))
        finally:
            self.data_collected_signal.emit(self.speed_data, self.time_stamps)
            self.finished_signal.emit()
    def reconnect(self):
        if self.RTK_singal == 0 and self.point == 0:
            while True:
                self.s = socket.socket()
                self.s.settimeout(10)
                self.s.connect((self.Ntrip_host, self.Ntrip_port))
                self.ntrip_request = f"GET /{self.Ntrip_mountpoint} HTTP/1.0\r\n"
                self.ntrip_request += "User-Agent: NTRIP PythonClient/1.0\r\n"
                self.ntrip_request += f"Authorization: Basic {base64.b64encode((self.Ntrip_user + ':' + self.Ntrip_password).encode()).decode()}\r\n"
                self.ntrip_request += "\r\n"
                self.s.send(self.ntrip_request.encode())
                response = self.s.recv(1024)
                if b"ICY 200 OK" in response:
                    self.RTK_singal = 1
                    break
        elif self.RTK_singal == 0 and self.point != 0:
            self.s = socket.socket()
            self.s.settimeout(10)
            self.s.connect((self.Ntrip_host, self.Ntrip_port))
            self.ntrip_request = f"GET /{self.Ntrip_mountpoint} HTTP/1.0\r\n"
            self.ntrip_request += "User-Agent: NTRIP PythonClient/1.0\r\n"
            self.ntrip_request += f"Authorization: Basic {base64.b64encode((self.Ntrip_user + ':' + self.Ntrip_password).encode()).decode()}\r\n"
            self.ntrip_request += "\r\n"
            self.s.send(self.ntrip_request.encode())
            response = self.s.recv(1024)
            if b"ICY 200 OK" in response:
                self.RTK_singal = 1
class Worker_radar(QThread):
    finished_signal = pyqtSignal()
    mistake_message_transmit = pyqtSignal(str)
    data_collected_signal = pyqtSignal(list, list, list, list, list)
    acc_data_collected_signal = pyqtSignal(list, list, list, list)
    gyr_data_collected_signal = pyqtSignal(list, list, list, list)
    ord_data_collected_signal = pyqtSignal(list, list, list, list)
    avg_data_parsed = pyqtSignal(float, float, float, float)
    def __init__(self, port, baudrate):
        super(Worker_radar, self).__init__()
        self.should_run = True
        self.port = port
        self.baudrate = baudrate
        print(self.port)
        print(self.baudrate)
        self.serial_connection = None
        self.speed_list = []
        self.angle_list = []
        self.force1_list = []
        self.force2_list = []
        self.accX_list = []
        self.accY_list = []
        self.accZ_list = []
        self.gyroX_list = []
        self.gyroY_list = []
        self.gyroZ_list = []
        self.roll_list = []
        self.pitch_list = []
        self.yaw_list = []
        self.time_stamps = []
        self.speed_data_for_avg = []
        self.angle_data_for_avg = []
        self.force1_data_for_avg = []
        self.force2_data_for_avg = []
    def run(self):
        try:
            self.serial_connection = serial.Serial(self.port, self.baudrate, timeout=1)
            while self.should_run:
                line = self.serial_connection.readline().decode('utf-8', errors='ignore').strip()
                if "Radar:" in line and "Yaw:" in line:
                    try:
                        data_pattern = r'Radar:(.*?),Angle:(.*?)°,Force1:(.*?)N,Force2:(.*?)N,AccX:(.*?)g,AccY:(.*?)g,AccZ:(.*?)g,GyroX:(.*?)°/s,GyroY:(.*?)°/s,GyroZ:(.*?)°/s,Roll:(.*?)°Pitch:(.*?)°Yaw:(.*?)°'
                        match = re.search(data_pattern, line)
                        if match:
                            radar_data, angle_data, force1_data, force2_data, accX, accY, accZ, gyroX, gyroY, gyroZ, roll, pitch, yaw = match.groups()
                            try:
                                radar_data = float(radar_data)
                                if math.isinf(radar_data):
                                    radar_data = 0.0
                            except ValueError:
                                radar_data = 0.0
                            print(angle_data)
                            print(accX)
                            angle_data = float(angle_data)
                            force1_data = float(force1_data)
                            force2_data = float(force2_data)
                            accX_data = float(accX)
                            accY_data = float(accY)
                            accZ_data = float(accZ)
                            gyroX_data = float(gyroX)
                            gyroY_data = float(gyroY)
                            gyroZ_data = float(gyroZ)
                            roll_data = float(roll)
                            pitch_data = float(pitch)
                            yaw_data = float(yaw)
                            self.speed_list.append(radar_data)
                            self.angle_list.append(angle_data)
                            self.force1_list.append(force1_data)
                            self.force2_list.append(force2_data)
                            self.accX_list.append(accX_data)
                            self.accY_list.append(accY_data)
                            self.accZ_list.append(accZ_data)
                            self.gyroX_list.append(gyroX_data)
                            self.gyroY_list.append(gyroY_data)
                            self.gyroZ_list.append(gyroZ_data)
                            self.roll_list.append(roll_data)
                            self.pitch_list.append(pitch_data)
                            self.yaw_list.append(yaw_data)
                            self.time_stamps.append(datetime.now())
                            print(accX_data)
                            self.speed_data_for_avg.append(radar_data)
                            self.angle_data_for_avg.append(angle_data)
                            self.force1_data_for_avg.append(force1_data)
                            self.force2_data_for_avg.append(force2_data)
                            if len(self.speed_data_for_avg) == 5:
                                avg_speed = sum(self.speed_data_for_avg) / 5
                                avg_angle = sum(self.angle_data_for_avg) / 5
                                avg_force1 = sum(self.force1_data_for_avg) / 5
                                avg_force2 = sum(self.force2_data_for_avg) / 5
                                self.avg_data_parsed.emit(avg_speed, avg_angle, avg_force1, avg_force2)
                                self.speed_data_for_avg.clear()
                                self.angle_data_for_avg.clear()
                                self.force1_data_for_avg.clear()
                                self.force2_data_for_avg.clear()
                    except Exception as err_1:
                        print(err_1)
                        self.mistake_message_transmit.emit(str(err_1))
                        continue
            self.finished_signal.emit()
        except Exception as err_1:
            self.mistake_message_transmit.emit(err_1)
        finally:
            self.data_collected_signal.emit(self.speed_list, self.angle_list, self.force1_list, self.force2_list,
                                            self.time_stamps)
            self.acc_data_collected_signal.emit(self.accX_list, self.accY_list, self.accZ_list, self.time_stamps)
            self.gyr_data_collected_signal.emit(self.gyroX_list, self.gyroY_list, self.gyroZ_list, self.time_stamps)
            self.ord_data_collected_signal.emit(self.roll_list, self.pitch_list, self.yaw_list, self.time_stamps)
            self.finished_signal.emit()
if __name__ == '__main__':
    try:
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
        app = QApplication(sys.argv)
        stacked_widget = QStackedWidget()
        Welcome_Window = WelcomeWindow(stacked_widget)
        information_Show_Window = information_show_Window(stacked_widget)
        stacked_widget.addWidget(Welcome_Window)
        stacked_widget.addWidget(information_Show_Window)
        stacked_widget.setCurrentIndex(0)
        main_window = QMainWindow()
        main_window.setWindowTitle("果园作业机械研究团队")
        main_window.setCentralWidget(stacked_widget)
        main_window.show()
        sys.exit(app.exec_())
    except Exception as err:
        print(err)
